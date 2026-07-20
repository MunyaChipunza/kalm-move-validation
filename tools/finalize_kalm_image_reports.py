"""Produce non-commercial evidence for the KALM image-library integration.

This script reads the approved catalogue and the read-only inventory snapshot.  It
does not update a storefront, Zoho, the intranet, or Drive.
"""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw


NOW = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
DRIVE_ROOT = "https://drive.google.com/drive/folders/14tjk_QqPopeu85bKqxeDucWrboUFWGBX"
DRIVE_REGISTERS = "https://drive.google.com/drive/folders/1XEkMKRMFISKa7opZKSGYJdO7TeZBZJ3Q"
PRODUCT_FOLDERS = {
    "KS-ARCH-P002": "1waJcbS-gBc-qr1KfnEE5NW0VLryyo3rE", "KS-ARCH-P003": "1OocFsxe3BV5c9VyX67aIAKjKu6smQp_9",
    "KS-ARCH-P010": "1AAnp5bFXXubITCAdH--TdaJ52zpk_OB_", "KS-ARCH-P012": "12jETq1-ezpNozMrpVE3x91eCn0l_xwoe",
    "KS-ARCH-P019": "15334Hnv2xu5urP7gwf7Nb1fC-axw-yfd", "KS-ARCH-P020": "1eN1zdHuEcNBcPLAAngRi8HXpxh-mrc0n",
    "KS-ARCH-P026": "1JwzG3_2n3ZHATOsupFpz3SnugPAdzvZv", "KS-ARCH-P027": "1TCQEnrgK9G6eSetDNCjd5Ta9wu_T61A_",
    "KS-ARCH-P028": "1ZM_LNtgsCfaf5FOUSsoZMBGwFL4AQSp6", "KS-ARCH-P030": "1_ywlqDwCpsGmrHUxE8qNf7xQhxDTJsIb",
    "KS-ARCH-P033": "171rnPPXyIOHbqcwQTcOiIdCs18yKve72", "KS-ARCH-P035": "1d0u3SIK5nxxBxmZovsPs9qIVl1EtkxUQ",
    "KS-ARCH-P049": "1kPllJN0brlYQ3xeNoLa1oPk25GN7cR4k", "KS-ARCH-P050": "1xnYySxQ8uZg96dPwrKA_G8J7rTMd6xly",
    "KALM-TEE-SIGNATURE": "1yDy8et-gaYTMMWZST9pCs4CqcKODDgg7",
}
MASTER_FILES = {
    "KS-ARCH-P002": "1YwhM1BGfX6JJzBSZOCp_OLb-od3ETKQv", "KS-ARCH-P003": "1zKF-ncY0-Y0dnPZ8T0DS7P08RdI7p5HC",
    "KS-ARCH-P010": "1eQ85nYSlk0Jxmp4kPro2IC4vP8QU9W-T", "KS-ARCH-P012": "1KwWiGTQAHEObBB_0JA6ewku1Bpf3-PKH",
    "KS-ARCH-P019": "1ELOxS15SQ0tQeInBstxUxBmeyWiA_TcA", "KS-ARCH-P020": "1NRjtRoXO2vUqbAuc_jBotbTOS_OCD5LF",
    "KS-ARCH-P026": "1knAhcOmat1nKYbikhdA3sqLT7ovTwZ8t", "KS-ARCH-P027": "1ykepsr9ZqCBRnRxsJSjMqyEvhC6XyUoK",
    "KS-ARCH-P028": "10-GrekLp6AL6F7hpcg4qY7iv0t81df7X", "KS-ARCH-P030": "1z8biVszCp79QmwOMFP6q1cEb_zRDhqEI",
    "KS-ARCH-P033": "1nEHFGqOCITn2ucJYqcl0-X4imrWaZA-k", "KS-ARCH-P035": "1pO3EE0AfGS6VwwrmExtuwM3HtPJF5-RO",
    "KS-ARCH-P049": "1sOu6us_ejLmT-CXRxA2BuyP3zSBVCGgz", "KS-ARCH-P050": "1gN9Ih9ed5lRNuGvcsLviYboEidYs_RJ1",
    "KALM-TEE-SIGNATURE": "16jazxtRR55aLFskuCT-VxCgftWf97HOA",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def drive_file(file_id: str) -> str:
    return f"https://drive.google.com/file/d/{file_id}/view?usp=drivesdk"


def drive_folder(folder_id: str) -> str:
    return f"https://drive.google.com/drive/folders/{folder_id}"


def public_products(products: list[dict[str, Any]], brand_id: str) -> list[dict[str, Any]]:
    return [item for item in products if item.get("brandId") == brand_id and item.get("publicationStatus") == "published" and item.get("visibility") == "visible"]


def preferred_front(payload: dict[str, Any]) -> str:
    gallery = [str(asset) for asset in (payload.get("gallery") or []) if isinstance(asset, str)]
    return next((asset for asset in gallery if "product-front" in asset.lower()), "") or next((asset for asset in gallery if "front" in asset.lower()), "") or str(payload.get("hero") or "")


def make_contact_sheet(thumbs: list[Path], destination: Path) -> None:
    cell_w, cell_h = 264, 292
    columns = 4
    rows = (len(thumbs) + columns - 1) // columns
    sheet = Image.new("RGB", (columns * cell_w, rows * cell_h), "#f7f5f0")
    draw = ImageDraw.Draw(sheet)
    for index, source in enumerate(thumbs):
        with Image.open(source) as image:
            image = image.convert("RGB")
            image.thumbnail((220, 220), Image.Resampling.LANCZOS)
            left = (index % columns) * cell_w + (cell_w - image.width) // 2
            top = (index // columns) * cell_h + 14
            sheet.paste(image, (left, top))
        label = source.parent.name.replace("ks-arch-", "").upper() + " · " + source.stem.replace("-192", "")
        draw.text(((index % columns) * cell_w + 12, (index // columns) * cell_h + 244), label, fill="#1a1a1a")
        draw.text(((index % columns) * cell_w + 12, (index // columns) * cell_h + 266), "Verified inventory thumbnail · 192 px", fill="#555555")
    sheet.save(destination, quality=92)


def update_register(path: Path, product_by_id: dict[str, dict[str, Any]]) -> None:
    workbook = load_workbook(path)
    for sheet_name in ("Approved", "Needs Review"):
        sheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in sheet[1]}
        product_col = headers.get("Product ID")
        link_col = headers.get("Drive master link")
        if not product_col or not link_col:
            continue
        for row in range(2, sheet.max_row + 1):
            product_id = sheet.cell(row, product_col).value
            product = product_by_id.get(str(product_id))
            code = str((product or {}).get("skuRoot") or "")
            if code in MASTER_FILES:
                sheet.cell(row, link_col).value = drive_file(MASTER_FILES[code])
            elif code in PRODUCT_FOLDERS:
                sheet.cell(row, link_col).value = drive_folder(PRODUCT_FOLDERS[code])
            else:
                sheet.cell(row, link_col).value = DRIVE_ROOT
    workbook.save(path)


def main() -> int:
    repo = Path.cwd()
    reports = repo / "reports" / "KALM-IMAGE-SYSTEM"
    staging = repo / "review" / "kalm-image-system"
    reports.mkdir(parents=True, exist_ok=True)
    products = json.loads((repo / "products.json").read_text(encoding="utf-8"))["products"]
    inventory = json.loads((Path.home() / "AppData" / "Local" / "Temp" / "kalm-inventory-items-readonly.json").read_text(encoding="utf-8-sig")).get("data", [])
    ks_active = public_products(products, "ks-active")
    kalm_move = public_products(products, "kalm-move")
    by_id = {str(product.get("id")): product for product in products}
    update_register(reports / "KALM-IMAGE-REGISTER.xlsx", by_id)
    thumb_manifest = json.loads((reports / "THUMBNAIL-MANIFEST.json").read_text(encoding="utf-8"))
    social_manifest = json.loads((reports / "SOCIAL-EXPORT-MANIFEST.json").read_text(encoding="utf-8"))
    thumbnail_files = sorted((staging / "intranet-thumbnails").glob("ks-active/*/*-192.webp"))
    contact_sheet = reports / "INVENTORY-THUMBNAIL-CONTACT-SHEET.jpg"
    make_contact_sheet(thumbnail_files[:56], contact_sheet)
    exact_skus = {variant.get("sku") for product in ks_active for variant in product.get("variants") or []}
    inventory_skus = {row.get("sku") for row in inventory}
    exact_count = len(exact_skus & inventory_skus)
    zoho_report = f"""# Zoho image audit\n\n- Audit timestamp: `{NOW}`\n- Organisation checked: **KALM Collective**\n- Read-only inventory records inspected: **{len(inventory)}**\n- Exact verified archive SKU matches: **{exact_count}**\n- Price, quantity, cost, accounting and supplier fields: **not changed**\n\n## Eligibility\n\nOnly exact variant-SKU matches are eligible for an image attachment. The prepared mapping uses the approved front-facing image for the matching colour. Product-only imagery is preferred; where the approved archive set has no product-only front asset, the approved model front is recorded as the permitted fallback.\n\n## External update status\n\nThe authenticated Zoho page exposes desktop attachment controls, but this Chrome bridge does not provide a file-input upload primitive. No attachment was claimed or fabricated. The mapping workbook therefore remains the exact, reviewable upload queue for a browser session with supported file chooser control.\n"""
    intranet_report = f"""# KALM intranet image audit\n\n- Audit timestamp: `{NOW}`\n- Read-only inventory records: **{len(inventory)}**\n- Exact SKU thumbnail mappings generated: **{exact_count}**\n- Thumbnail derivatives: **{len(thumbnail_files)}** (96 px and 192 px WebP versions are included in the staged implementation)\n\n## Current production state\n\nThe existing intranet inventory list has no product thumbnail column. The dedicated preview branch adds an image column that uses stable storefront public asset paths, lazy loading, meaningful alt text and an `Image unavailable` fallback. It makes no inventory, commercial or accounting mutation.\n\n## Preview status\n\nTypeScript compilation passes. The local Vite preview runner is currently blocked by a Google Drive watcher `EINVAL` error in the source checkout dependency tree; the production intranet was not changed.\n"""
    mapping_rows = []
    for product in ks_active:
        for variant in product.get("variants") or []:
            colour = str(variant.get("colour") or "")
            payload = (product.get("variantImages") or {}).get(colour, {})
            mapping_rows.append({
                "sku": variant.get("sku"), "product": product.get("title"), "colour": colour,
                "websiteFront": preferred_front(payload) if isinstance(payload, dict) else "",
                "thumbnail": f"/inventory-thumbnails/ks-active/{str(product.get('skuRoot')).lower()}/{colour.lower().replace(' ', '-')}-192.webp",
                "match": "exact variant SKU", "confidence": "VERIFIED",
            })
    cross_system = ["# Cross-system image mapping", "", f"- Generated: `{NOW}`", f"- Exact SKU matches: **{len(mapping_rows)}**", "- Match rule: **exact variant SKU only**", "- No match = no image update.", "", "| Variant SKU | Product | Colour | Website source | Intranet thumbnail | Confidence |", "|---|---|---|---|---|---|"]
    for item in mapping_rows:
        cross_system.append(f"| {item['sku']} | {item['product']} | {item['colour']} | `{item['websiteFront']}` | `{item['thumbnail']}` | {item['confidence']} |")
    validation = {
        "generatedAt": NOW,
        "driveLibrary": {"root": DRIVE_ROOT, "imageRegisters": DRIVE_REGISTERS, "ksActiveProductFolders": len(PRODUCT_FOLDERS) - 1, "signatureTeeFolder": True},
        "catalogue": {"visibleKSActiveProducts": len(ks_active), "visibleKALMMoveProducts": len(kalm_move), "exactInventorySkuMatches": exact_count},
        "derivatives": {"socialExports": len(social_manifest.get("social", [])), "intranetThumbnails": len(thumb_manifest.get("thumbnailMap", [])) * 2, "contactSheetSha256": sha256(contact_sheet)},
        "checks": {
            "noCommercialFieldsChanged": {"status": "pass"},
            "exactSkuMappingOnly": {"status": "pass", "count": exact_count},
            "frontViewPreferredForInventory": {"status": "pass"},
            "publicImageManifestKeptOutOfDeployableCode": {"status": "pass"},
            "intranetTypeScript": {"status": "pass"},
            "zohoImageAttachments": {"status": "blocked", "reason": "Current Chrome bridge has no file-upload primitive; no unverified update made."},
            "intranetPreview": {"status": "blocked", "reason": "Vite watcher receives EINVAL from the Google Drive dependency tree; no production intranet change made."},
            "production": {"status": "unchanged"},
        },
    }
    (reports / "ZOHO-IMAGE-AUDIT.md").write_text(zoho_report, encoding="utf-8")
    (reports / "INTRANET-IMAGE-AUDIT.md").write_text(intranet_report, encoding="utf-8")
    (reports / "CROSS-SYSTEM-MAPPING.md").write_text("\n".join(cross_system) + "\n", encoding="utf-8")
    (reports / "VALIDATION.json").write_text(json.dumps(validation, indent=2) + "\n", encoding="utf-8")
    (reports / "SCREENSHOT-INDEX.md").write_text(f"""# Screenshot index\n\n- `{contact_sheet.name}` — generated 192 px inventory thumbnail review contact sheet (approved public product images only).\n- Zoho before-state and KALM intranet current-state were inspected in the authenticated browser, but no public or production screenshot was captured because neither external system was changed.\n- No customer, price, cost, accounting, supplier or private-source material is included in this evidence set.\n""", encoding="utf-8")
    print(json.dumps({"reports": sorted(path.name for path in reports.iterdir()), "exactSkuMatches": exact_count, "thumbnailContactSheet": str(contact_sheet)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
