#!/usr/bin/env python3
"""Build the KALM image-system audit, non-public library staging, and registers.

This script is deliberately catalogue-led: only assets referenced by an active,
approved product are selected for the reusable library, social formats, and
inventory thumbnails. It never changes products.json or any commercial data.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff", ".svg"}
PRODUCT_FIELDS = (
    "id", "title", "brand", "brandId", "skuRoot", "sku", "slug", "audience",
    "publicationStatus", "visibility", "image", "gallery", "variantImages", "colors", "variants",
)
APPROVED_ARCHIVE = "assets/images/products/ks-active/archive-approved/"
NOW = datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalise(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")


def safe_name(value: str | None) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", value or "unknown").strip("-").upper()


def detect_view(path: str) -> str:
    lower = Path(path).stem.lower()
    for view in ("three-quarter", "three_quarter", "back", "side", "front", "detail", "flat-lay", "flat_lay", "hero"):
        if view in lower:
            return view.replace("_", "-")
    return "product"


def dimensions(path: Path) -> tuple[int | None, int | None, str | None]:
    if path.suffix.lower() == ".svg":
        text = path.read_text(encoding="utf-8", errors="ignore")[:4096]
        width = re.search(r"\bwidth=[\"']?(\d+)", text)
        height = re.search(r"\bheight=[\"']?(\d+)", text)
        return (int(width.group(1)) if width else None, int(height.group(1)) if height else None, "SVG")
    try:
        with Image.open(path) as image:
            return image.width, image.height, image.format
    except Exception:
        return None, None, None


def classify_path(relative: str, referenced: bool) -> tuple[str, str, str]:
    lower = relative.lower().replace("\\", "/")
    if any(token in lower for token in ("review-only", "/review/", "source-reference", "private", "kuhle")):
        return "internal or private evidence", "restricted", "not for public reuse"
    if any(token in lower for token in ("rejected", "quarantine", "failed", "cad-like", "bottles-v2", "bottles-v3")):
        return "rejected", "rejected", "do not reuse"
    if any(token in lower for token in ("supplier", "alibaba", "reference", "staged")):
        return "supplier or source reference", "uncertain", "not for public reuse"
    if "campaign" in lower:
        return "campaign", "approved" if referenced else "needs review", "campaign or social suitability review"
    if referenced and APPROVED_ARCHIVE in lower:
        return "approved ecommerce", "approved", "approved public Archive asset"
    if referenced:
        return "ecommerce", "approved", "currently catalogued asset"
    return "unclassified", "needs review", "not currently referenced by the visible catalogue"


def inspect_image(path: Path, repo: Path, ref_map: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    relative = path.relative_to(repo).as_posix()
    refs = ref_map.get(relative, [])
    classification, approval, notes = classify_path(relative, bool(refs))
    width, height, image_format = dimensions(path)
    return {
        "relativePath": relative, "filename": path.name, "extension": path.suffix.lower(), "bytes": path.stat().st_size,
        "width": width, "height": height, "format": image_format, "sha256": sha256(path), "classification": classification,
        "approvalStatus": approval, "notes": notes, "referencedBy": refs,
    }


def image_refs(product: dict[str, Any]) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    if isinstance(product.get("image"), str):
        refs.append({"path": product["image"], "colour": "", "role": "hero"})
    for path in product.get("gallery") or []:
        if isinstance(path, str):
            refs.append({"path": path, "colour": "", "role": "gallery"})
    for colour, payload in (product.get("variantImages") or {}).items():
        if not isinstance(payload, dict):
            continue
        if isinstance(payload.get("hero"), str):
            refs.append({"path": payload["hero"], "colour": colour, "role": "variant hero"})
        for path in payload.get("gallery") or []:
            if isinstance(path, str):
                refs.append({"path": path, "colour": colour, "role": "variant gallery"})
    unique: dict[tuple[str, str], dict[str, str]] = {}
    for entry in refs:
        unique.setdefault((entry["path"], entry["colour"]), entry)
    return list(unique.values())


def visible_product(product: dict[str, Any]) -> bool:
    return product.get("publicationStatus") == "published" and product.get("visibility") == "visible"


def selected_products(products: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        product for product in products
        if visible_product(product) and product.get("brandId") in {"ks-active", "kalm-move"}
    ]


def product_code(product: dict[str, Any]) -> str:
    return str(product.get("skuRoot") or product.get("sku") or product.get("id") or "UNKNOWN")


def library_bucket(product: dict[str, Any]) -> str:
    if product.get("id") == "KALM-TEE-SIGNATURE-001":
        return "KALM Collective"
    return "KS Active" if product.get("brandId") == "ks-active" else "KALM Move"


def create_canvas(image: Image.Image, size: tuple[int, int]) -> Image.Image:
    """Fit without crop/distortion on a neutral ecommerce-safe canvas."""
    source = image.convert("RGB")
    contained = ImageOps.contain(source, size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", size, "#f6f4ef")
    canvas.paste(contained, ((size[0] - contained.width) // 2, (size[1] - contained.height) // 2))
    return canvas


def render_social(source: Path, destination: Path, size: tuple[int, int]) -> bool:
    try:
        with Image.open(source) as image:
            out = create_canvas(image, size)
            destination.parent.mkdir(parents=True, exist_ok=True)
            out.save(destination, "JPEG", quality=92, optimize=True, progressive=True)
        return True
    except Exception:
        return False


def render_thumbnail(source: Path, destination: Path, size: int) -> bool:
    try:
        with Image.open(source) as image:
            out = create_canvas(image, (size, size))
            destination.parent.mkdir(parents=True, exist_ok=True)
            out.save(destination, "WEBP", quality=92, method=6)
        return True
    except Exception:
        return False


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_markdown(path: Path, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def set_sheet_style(sheet, freeze: str = "A2") -> None:
    navy = "17233A"
    gold = "C8A96B"
    pale = "F6F4EF"
    thin = Side(style="thin", color="D9D4C8")
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=gold))
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=pale)


def populate_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    set_sheet_style(sheet)
    widths = {
        "Image ID": 16, "Product Name": 32, "Master filename": 42, "Drive master link": 46,
        "Website URL": 56, "Notes": 46, "Source type": 26, "Rights or origin": 28,
        "Approval status": 18, "Product-truth status": 20, "Colour": 18, "SKU": 25,
        "Variant SKU": 30, "Instagram 4:5 link": 46, "Instagram 1:1 link": 46,
        "Story 9:16 link": 46, "Last reviewed": 24,
    }
    for index, name in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(name, 20)
    if rows:
        status_column = headers.index("Approval status") + 1 if "Approval status" in headers else None
        if status_column:
            letter = get_column_letter(status_column)
            sheet.conditional_formatting.add(
                f"{letter}2:{letter}{len(rows)+1}",
                FormulaRule(formula=[f'${letter}2="approved"'], fill=PatternFill("solid", fgColor="D9EAD3")),
            )
            sheet.conditional_formatting.add(
                f"{letter}2:{letter}{len(rows)+1}",
                FormulaRule(formula=[f'${letter}2="rejected"'], fill=PatternFill("solid", fgColor="F4CCCC")),
            )


def build_workbooks(output: Path, selected: list[dict[str, Any]], audit: list[dict[str, Any]], inventory_rows: list[dict[str, Any]], social: list[dict[str, Any]]) -> None:
    headers = [
        "Image ID", "Brand", "Parent Brand", "Product ID", "Product Name", "SKU", "Variant SKU", "Colour",
        "Size applicability", "Image role", "View", "Model or product-only", "Master filename", "Drive master link",
        "Ecommerce filename", "Website URL", "Instagram 4:5 link", "Instagram 1:1 link", "Story 9:16 link",
        "Approval status", "Product-truth status", "Source type", "Source date", "Rights or origin", "Notes", "Last reviewed",
    ]
    by_path = {row["relativePath"]: row for row in audit}
    social_by_source = defaultdict(dict)
    for row in social:
        social_by_source[row["source"]][row["format"]] = row["relativePath"]
    archive_rows: list[list[Any]] = []
    needs_rows: list[list[Any]] = []
    concepts_rows: list[list[Any]] = []
    rejected_rows: list[list[Any]] = []
    missing_rows: list[list[Any]] = []
    duplicate_rows: list[list[Any]] = []
    hashes = Counter(row["sha256"] for row in audit if row.get("sha256"))
    for product in selected:
        product_refs = image_refs(product)
        variants = product.get("variants") or []
        variant_by_colour = defaultdict(list)
        for variant in variants:
            variant_by_colour[variant.get("colour") or ""].append(variant)
        for ref in product_refs:
            record = by_path.get(ref["path"])
            if not record:
                missing_rows.append([
                    f"IMG-MISSING-{len(missing_rows)+1:04d}", product.get("brand"), "KALM Collective", product.get("id"), product.get("title"), product_code(product), "", ref.get("colour"), "", ref.get("role"), detect_view(ref["path"]), "unknown", Path(ref["path"]).name, "", "", "", "", "", "", "missing", "unverified", "missing", NOW[:10], "local catalog reference", "Referenced by product record but no local file exists.", NOW,
                ])
                continue
            sku_values = variant_by_colour.get(ref.get("colour") or "", [])
            sku = sku_values[0].get("sku") if sku_values else ""
            model = "model" if any(token in record["relativePath"].lower() for token in ("male", "female", "model", "back", "side")) else "product-only"
            row = [
                f"IMG-{record['sha256'][:12].upper()}", product.get("brand"), "KALM Collective", product.get("id"), product.get("title"), product_code(product), sku, ref.get("colour") or "", ", ".join(sorted({str(v.get('size') or '') for v in sku_values if v.get('size')})), ref.get("role"), detect_view(ref["path"]), model, Path(ref["path"]).name,
                "Drive library upload recorded separately", Path(ref["path"]).name,
                f"https://kalmcollective.co.za/{ref['path']}",
                social_by_source.get(ref["path"], {}).get("Instagram 4:5", ""),
                social_by_source.get(ref["path"], {}).get("Instagram 1:1", ""),
                social_by_source.get(ref["path"], {}).get("Story 9:16", ""),
                record["approvalStatus"], "verified" if record["approvalStatus"] == "approved" else "needs review", record["classification"], NOW[:10], "repository-approved catalogue asset" if record["approvalStatus"] == "approved" else "not cleared for social/public reuse", record["notes"], NOW,
            ]
            if record["approvalStatus"] == "approved":
                archive_rows.append(row)
            elif record["approvalStatus"] == "rejected":
                rejected_rows.append(row)
            elif "concept" in record["classification"]:
                concepts_rows.append(row)
            else:
                needs_rows.append(row)
            if record.get("sha256") and hashes[record["sha256"]] > 1:
                duplicate_rows.append(row + [hashes[record["sha256"]]])
    register = Workbook()
    register.remove(register.active)
    populate_sheet(register, "Approved", headers, archive_rows)
    populate_sheet(register, "Needs Review", headers, needs_rows)
    populate_sheet(register, "Concepts", headers, concepts_rows)
    populate_sheet(register, "Rejected", headers, rejected_rows)
    populate_sheet(register, "Missing Images", headers, missing_rows)
    populate_sheet(register, "Duplicate Images", headers + ["Duplicate Count"], duplicate_rows)
    register.save(output / "KALM-IMAGE-REGISTER.xlsx")

    mapping_headers = [
        "Brand", "Product Name", "Website Product ID", "Website SKU", "Variant SKU", "Colour", "Zoho Item ID", "Zoho SKU", "Intranet Item ID", "Intranet SKU", "Master Image ID", "Thumbnail Image ID", "Website URL", "Zoho image status", "Intranet image status", "Match method", "Match confidence", "Exception reason", "Review status",
    ]
    product_by_sku = {product_code(product): product for product in selected if product.get("brandId") == "ks-active"}
    rows: list[list[Any]] = []
    for inventory in inventory_rows:
        sku = inventory.get("sku") or ""
        code_match = re.search(r"KS-ARCH-(P\d{3})", sku)
        root = f"KS-ARCH-{code_match.group(1)}" if code_match else ""
        product = product_by_sku.get(root)
        colour = ""
        variant = None
        if product:
            variant = next((item for item in product.get("variants") or [] if item.get("sku") == sku), None)
            colour = (variant or {}).get("colour") or ""
        image_payload = ((product or {}).get("variantImages") or {}).get(colour, {})
        gallery = [str(item) for item in (image_payload.get("gallery") or []) if isinstance(item, str)] if isinstance(image_payload, dict) else []
        image = next((item for item in gallery if "product-front" in item.lower()), "") or next((item for item in gallery if "front" in item.lower()), "") or (image_payload.get("hero", "") if isinstance(image_payload, dict) else "")
        thumbnail = f"inventory-thumbnails/ks-active/{normalise(root)}/{normalise(colour)}-192.webp" if image else ""
        exact = bool(product and variant and sku == inventory.get("sku"))
        rows.append([
            "KS Active", (product or {}).get("title", inventory.get("item_name", "")), (product or {}).get("id", ""), root, sku, colour,
            inventory.get("item_id", ""), sku, inventory.get("item_id", ""), sku,
            f"IMG-{by_path.get(image, {}).get('sha256', '')[:12].upper()}" if image else "", f"THUMB-{normalise(root)}-{normalise(colour)}-192" if thumbnail else "",
            f"https://kalmcollective.co.za/{image}" if image else "", "pending exact verified upload" if exact else "not eligible",
            "preview mapped" if exact else "not eligible", "exact variant SKU" if exact else "no exact local variant", "VERIFIED" if exact else "UNRESOLVED", "" if exact else "Website product or variant mapping missing", "READY_FOR_IMAGE_UPDATE" if exact else "REVIEW_REQUIRED",
        ])
    mapping = Workbook()
    mapping.remove(mapping.active)
    populate_sheet(mapping, "SKU Mapping", mapping_headers, rows)
    mapping.save(output / "KALM-IMAGE-SKU-MAPPING.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument("--inventory-json", type=Path, required=True)
    args = parser.parse_args()
    repo = args.repo.resolve()
    reports = repo / "reports" / "KALM-IMAGE-SYSTEM"
    staging = repo / "review" / "kalm-image-system"
    thumbs = staging / "intranet-thumbnails"
    social_root = staging / "social"
    products = json.loads((repo / "products.json").read_text(encoding="utf-8"))["products"]
    selected = selected_products(products)
    ref_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        for ref in image_refs(product):
            ref_map[ref["path"]].append({"productId": product.get("id"), "colour": ref["colour"], "role": ref["role"], "visible": visible_product(product)})
    audit: list[dict[str, Any]] = []
    ignored_parts = {".git", "node_modules", ".netlify", ".netlify-runtime", "dist", "tmp", "__pycache__"}
    candidates = [
        path for path in repo.rglob("*")
        if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES and not any(part in ignored_parts for part in path.parts)
    ]
    with ThreadPoolExecutor(max_workers=12) as pool:
        audit = list(pool.map(lambda item: inspect_image(item, repo, ref_map), candidates))
    audit.sort(key=lambda row: row["relativePath"])
    inventory_payload = json.loads(args.inventory_json.read_text(encoding="utf-8-sig"))
    inventory_rows = inventory_payload.get("data") if isinstance(inventory_payload, dict) else []
    inventory_rows = inventory_rows or []
    public_ks = [product for product in selected if product.get("brandId") == "ks-active"]
    public_move = [product for product in selected if product.get("brandId") == "kalm-move"]
    lookup = {row["relativePath"]: row for row in audit}
    social: list[dict[str, str]] = []
    selected_primaries = []
    for product in selected:
        if product.get("image"):
            selected_primaries.append((product, str(product["image"]), ""))
        if product.get("id") == "KALM-TEE-SIGNATURE-001":
            for colour, payload in (product.get("variantImages") or {}).items():
                if isinstance(payload, dict) and payload.get("hero"):
                    selected_primaries.append((product, str(payload["hero"]), str(colour)))
    seen_sources: set[tuple[str, str]] = set()
    for product, relative, colour in selected_primaries:
        key = (relative, colour)
        if key in seen_sources:
            continue
        seen_sources.add(key)
        record = lookup.get(relative)
        if not record or record["approvalStatus"] != "approved":
            continue
        source = repo / relative
        identity = "-".join(filter(None, [safe_name(product.get("brand")), safe_name(product_code(product)), safe_name(product.get("title")), safe_name(colour or "DEFAULT"), safe_name(detect_view(relative))]))
        for label, dimensions_target, suffix in (
            ("Instagram 4:5", (1080, 1350), "IG45"),
            ("Instagram 1:1", (1080, 1080), "IG11"),
            ("Story 9:16", (1080, 1920), "STORY916"),
        ):
            destination = social_root / normalise(product.get("brand")) / f"{identity}-{suffix}-V1.jpg"
            if render_social(source, destination, dimensions_target):
                social.append({"productId": str(product.get("id")), "source": relative, "format": label, "relativePath": destination.relative_to(repo).as_posix(), "file": str(destination)})
    def thumbnail_source(payload: dict[str, Any]) -> str:
        """Prefer a front-facing product view; use the approved hero only as a fallback."""
        gallery = [str(item) for item in (payload.get("gallery") or []) if isinstance(item, str)]
        for marker in ("product-front", "front", "flat-lay", "flat_lay"):
            candidate = next((item for item in gallery if marker in item.lower()), "")
            if candidate:
                return candidate
        return str(payload.get("hero") or "")

    thumbnail_count = 0
    thumbnail_map: list[dict[str, str]] = []
    for product in public_ks:
        root = product_code(product)
        for colour, payload in (product.get("variantImages") or {}).items():
            if not isinstance(payload, dict):
                continue
            preferred_source = thumbnail_source(payload)
            if not preferred_source:
                continue
            source = repo / preferred_source
            if not source.exists():
                continue
            folder = thumbs / "ks-active" / normalise(root)
            generated = []
            for size in (96, 192):
                destination = folder / f"{normalise(colour)}-{size}.webp"
                if render_thumbnail(source, destination, size):
                    thumbnail_count += 1
                    generated.append(destination.relative_to(repo).as_posix())
            if generated:
                thumbnail_map.append({"skuRoot": root, "colour": colour, "source": preferred_source, "thumbnails": generated})
    counts = Counter(row["classification"] for row in audit)
    approval_counts = Counter(row["approvalStatus"] for row in audit)
    audit_summary = {
        "generatedAt": NOW, "repo": str(repo), "scannedImageCount": len(audit), "classificationCounts": dict(counts),
        "approvalCounts": dict(approval_counts), "visibleKSActiveProducts": len(public_ks), "visibleKALMMoveProducts": len(public_move),
        "selectedSocialExports": len(social), "generatedIntranetThumbnails": thumbnail_count, "inventoryRecordCount": len(inventory_rows),
        "publicAssetManifestLocation": "Google Drive Image Registers (private; not stored in deployable website code)",
    }
    write_json(staging / "library-manifest-private.json", {"summary": audit_summary, "audit": audit, "social": social, "thumbnailMap": thumbnail_map})
    write_json(reports / "IMAGE-SOURCE-AUDIT.json", audit_summary)
    lines = [
        "# KALM Image Source Audit", "", f"Generated: `{NOW}`", "", "## Scope", "",
        f"- Repository images scanned: **{len(audit)}**", f"- Visible approved KS Active Archive products: **{len(public_ks)}**", f"- Visible KALM Move products: **{len(public_move)}**", f"- Existing Zoho-backed inventory records read: **{len(inventory_rows)}**", f"- Social derivatives generated without cropping or distortion: **{len(social)}**", f"- Dedicated 96 px / 192 px intranet thumbnails generated: **{thumbnail_count}**", "", "## Classification", "",
    ]
    for label, count in sorted(counts.items()):
        lines.append(f"- {label}: {count}")
    lines.extend(["", "## Approval state", ""])
    for label, count in sorted(approval_counts.items()):
        lines.append(f"- {label}: {count}")
    lines.extend([
        "", "## Controls", "",
        "- This audit is catalogue-led and does not mutate products, inventory, prices, costs, or accounting records.",
        "- Supplier, private, review-only, staged, and rejected sources are classified separately and are excluded from public reuse.",
        "- The complete public image manifest is staged only for the private Drive Image Registers folder; no Drive source links are exposed in deployable storefront code.",
        "- Social exports preserve the full source composition on neutral canvases; no logo, garment, or model is cropped or distorted.",
    ])
    write_markdown(reports / "IMAGE-SOURCE-AUDIT.md", lines)
    build_workbooks(reports, selected, audit, inventory_rows, social)
    write_json(reports / "THUMBNAIL-MANIFEST.json", {"generatedAt": NOW, "thumbnailMap": thumbnail_map, "count": thumbnail_count})
    write_json(reports / "SOCIAL-EXPORT-MANIFEST.json", {"generatedAt": NOW, "social": social, "count": len(social)})
    print(json.dumps(audit_summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
